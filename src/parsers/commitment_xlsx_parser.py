import os
import re
import openpyxl

from .common import clean_amount, extract_xlsx_report_date

DASH_ONLY_RE = re.compile(r'^-+$')
ORDER_RE = re.compile(r'^-{2,3}\s*ORDER NO\s*:\s*(\S+)\s+(\d+)\s+(\d+)-*$', re.IGNORECASE)
# A multi-line order sometimes continues on a bare "seq  gl_account----" row
# with no "--- ORDER NO :" prefix, since it belongs to the order above it.
ORDER_CONTINUATION_RE = re.compile(r'^(\d+)\s+(\d+)-*$')
VENDOR_RE = re.compile(r'^={2,}(.+?)-*$')
DETAIL_TOTAL_RE = re.compile(r'^DETAIL TOTAL-*$', re.IGNORECASE)


def parse_level_code(cell):
    """Return (letter, number, is_total) for a hierarchy code cell, or None.

    Handles two real-world quirks seen across weekly dumps:
    - 'TOTAL  I 003' style closing lines (double space before the letter).
    - the R-level letter sometimes dropped entirely, leaving a bare int
      (3 or 4) in the TYPE column instead of 'R 003' / 'R 004'.
    """
    if cell is None:
        return None
    if isinstance(cell, (int, float)):
        return ('R', int(cell), False)
    s = re.sub(r'\s+', ' ', str(cell).strip())
    m = re.match(r'^(TOTAL)?\s*([FRI])\s*(\d+)$', s)
    if not m:
        return None
    return (m.group(2), int(m.group(3)), bool(m.group(1)))


def locate_columns(ws):
    """Find the header row and the description/amount column indices.

    Column position shifts between weekly exports (5-column vs 6-column
    layouts), so columns are located by header text rather than hardcoded
    indices.
    """
    header_row_idx, header_vals = None, None
    for r in range(1, 11):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if any(isinstance(v, str) and 'DESCRIPTION' in v.upper() for v in vals):
            header_row_idx, header_vals = r, vals
            break
    if header_vals is None:
        raise ValueError(f"Could not locate header row in sheet {ws.title!r}")

    def find_col(*keywords):
        for i, v in enumerate(header_vals):
            if isinstance(v, str) and all(k in v.upper().replace(' ', '') for k in keywords):
                return i
        return None

    desc_col = find_col('DESCRIPTION')
    grv_col = find_col('GRV')
    post_col = find_col('POSTDATED')
    bal_col = find_col('COMMITMENT', 'BALANCE')
    return header_row_idx, desc_col, grv_col, post_col, bal_col


def extract_description(row_vals, desc_col, end_col):
    # The 6-column layout sometimes splits a long description mid-word across
    # the description column and its spacer overflow column (e.g. 'AUDIT
    # FEES:EXT PREVIOUS YEA' | 'R'), with no space in the source between
    # them, so the parts are concatenated directly rather than joined with a
    # separator.
    parts = [str(v) for v in row_vals[desc_col:end_col] if v not in (None, '')]
    return ''.join(parts).strip()


def parse_commitment_sheet(ws, source_file):
    header_row_idx, desc_col, grv_col, post_col, bal_col = locate_columns(ws)
    amount_start = min(c for c in (grv_col, post_col, bal_col) if c is not None)

    rows_out = []
    warnings = []
    ctx = {'fund': None, 'r1': None, 'r2': None, 'item_group': None, 'item': None}
    pending_orders = []
    last_order_no = None

    # Running subtotals mirror the report's own nested TOTAL lines, so we can
    # validate our reconstruction against the mainframe's printed figures.
    item_group_sum = [0.0, 0.0, 0.0]
    r2_sum = [0.0, 0.0, 0.0]
    r1_sum = [0.0, 0.0, 0.0]

    def add_to_running(grv, post, bal):
        for acc in (item_group_sum, r2_sum, r1_sum):
            acc[0] += grv
            acc[1] += post
            acc[2] += bal

    def make_row(vendor, order_no, order_seq, gl_account, grv, post, bal):
        return {
            'source_file': source_file,
            'sheet': ws.title,
            'fund_code': ctx['fund'][0] if ctx['fund'] else None,
            'fund_desc': ctx['fund'][1] if ctx['fund'] else None,
            'resp1_code': ctx['r1'][0] if ctx['r1'] else None,
            'resp1_desc': ctx['r1'][1] if ctx['r1'] else None,
            'resp2_code': ctx['r2'][0] if ctx['r2'] else None,
            'resp2_desc': ctx['r2'][1] if ctx['r2'] else None,
            'item_group_code': ctx['item_group'][0] if ctx['item_group'] else None,
            'item_group_desc': ctx['item_group'][1] if ctx['item_group'] else None,
            'item_code': ctx['item'][0] if ctx['item'] else None,
            'item_desc': ctx['item'][1] if ctx['item'] else None,
            'order_no': order_no,
            'order_seq': order_seq,
            'gl_account': gl_account,
            'vendor': vendor,
            'unpaid_grvs': grv,
            'post_dated_payments': post,
            'commitments_balance': bal,
        }

    for r in range(header_row_idx + 2, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        a = row_vals[0]
        desc_text = extract_description(row_vals, desc_col, amount_start)
        grv_raw = row_vals[grv_col] if grv_col is not None else None
        post_raw = row_vals[post_col] if post_col is not None else None
        bal_raw = row_vals[bal_col] if bal_col is not None else None

        level = parse_level_code(a)
        if level is not None:
            letter, num, is_total = level
            if is_total:
                reported = (clean_amount(grv_raw), clean_amount(post_raw), clean_amount(bal_raw))
                if letter == 'I' and num == 3:
                    actual = tuple(item_group_sum)
                    label = f"I003 {ctx['r1'][1] if ctx['r1'] else '?'} / {ctx['r2'][1] if ctx['r2'] else '?'} / {ctx['item_group'][1] if ctx['item_group'] else '?'}"
                elif letter == 'R' and num == 4:
                    actual = tuple(r2_sum)
                    label = f"R004 {ctx['r2'][1] if ctx['r2'] else '?'}"
                elif letter == 'R' and num == 3:
                    actual = tuple(r1_sum)
                    label = f"R003 {ctx['r1'][1] if ctx['r1'] else '?'}"
                else:
                    actual = None
                    label = None
                if actual is not None and any(abs(a - b) > 0.01 for a, b in zip(actual, reported)):
                    warnings.append(f"[{ws.title}] TOTAL mismatch for {label}: computed={actual} reported={reported}")
                continue
            if letter == 'F':
                ctx.update(fund=(f'F{num:03d}', desc_text), r1=None, r2=None, item_group=None, item=None)
            elif letter == 'R' and num == 3:
                ctx.update(r1=(f'R{num:03d}', desc_text), r2=None, item_group=None, item=None)
                r1_sum[:] = [0.0, 0.0, 0.0]
            elif letter == 'R' and num == 4:
                ctx.update(r2=(f'R{num:03d}', desc_text), item_group=None, item=None)
                r2_sum[:] = [0.0, 0.0, 0.0]
            elif letter == 'I' and num == 3:
                ctx.update(item_group=(f'I{num:03d}', desc_text), item=None)
                pending_orders = []
                item_group_sum[:] = [0.0, 0.0, 0.0]
            elif letter == 'I':
                ctx.update(item=(f'I{num:03d}', desc_text))
                pending_orders = []
            continue

        if desc_text == '':
            if any(isinstance(v, str) and DASH_ONLY_RE.match(v.strip()) for v in (grv_raw, post_raw, bal_raw)):
                continue  # dash separator row
            if grv_raw is None and post_raw is None and bal_raw is None:
                continue  # blank row

        m_order = ORDER_RE.match(desc_text)
        if m_order:
            last_order_no = m_order.group(1)
            pending_orders.append({
                'order_no': m_order.group(1),
                'order_seq': m_order.group(2),
                'gl_account': m_order.group(3),
                'grv': clean_amount(grv_raw),
                'post': clean_amount(post_raw),
                'bal': clean_amount(bal_raw),
            })
            continue

        m_cont = ORDER_CONTINUATION_RE.match(desc_text)
        if m_cont:
            pending_orders.append({
                'order_no': last_order_no,
                'order_seq': m_cont.group(1),
                'gl_account': m_cont.group(2),
                'grv': clean_amount(grv_raw),
                'post': clean_amount(post_raw),
                'bal': clean_amount(bal_raw),
            })
            continue

        m_vendor = VENDOR_RE.match(desc_text)
        if m_vendor:
            vendor_name = m_vendor.group(1).strip()
            if pending_orders:
                for o in pending_orders:
                    rows_out.append(make_row(vendor_name, o['order_no'], o['order_seq'], o['gl_account'], o['grv'], o['post'], o['bal']))
                    add_to_running(o['grv'], o['post'], o['bal'])
                pending_orders = []
            else:
                grv, post, bal = clean_amount(grv_raw), clean_amount(post_raw), clean_amount(bal_raw)
                rows_out.append(make_row(vendor_name, None, None, None, grv, post, bal))
                add_to_running(grv, post, bal)
            continue

        if DETAIL_TOTAL_RE.match(desc_text):
            pending_orders = []
            continue

        # Any other row shape (shouldn't occur) is silently skipped for this POC.

    return rows_out, warnings


def parse_commitment_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    report_date = extract_xlsx_report_date(wb)
    source_file = os.path.basename(path)
    all_rows, all_warnings = [], []
    for sheet_name in wb.sheetnames:
        rows, warnings = parse_commitment_sheet(wb[sheet_name], source_file=source_file)
        for row in rows:
            row['report_date'] = report_date
        all_rows.extend(rows)
        all_warnings.extend(warnings)
    return all_rows, all_warnings
