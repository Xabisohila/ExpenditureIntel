import openpyxl
import sys
import glob
import os

folder = os.path.join(os.path.dirname(__file__), "..", "data", "raw")
files = sorted(glob.glob(os.path.join(folder, "*.xlsx")))

for f in files:
    print("=" * 80)
    print(os.path.basename(f))
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
    except Exception as e:
        print("  ERROR loading:", e)
        continue
    print("  Sheets:", wb.sheetnames)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"  --- Sheet '{sheet_name}' dims: {ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column}")
        # print first 5 rows
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(6, ws.max_row), values_only=True)):
            print(f"    row{i+1}:", row)
